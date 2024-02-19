import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By  # Eksik olan import satırı
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
browser = webdriver.Chrome()


url = "https://www.msc.com/"
browser.get(url)    


element = browser.find_element(By.XPATH, "//*[@id='onetrust-accept-btn-handler']") .click() 
time.sleep(8)

tikla = browser.find_element(By.XPATH, "/html/body/div[1]/div[4]/div/div[2]/div/div[2]/div[1]/div[1]/div[2]/label")
tikla.click()
time.sleep(9)   

giris = browser.find_element(By.XPATH , "/html/body/div[1]/div[3]/div/div[2]/div/div[2]/div[1]/div[2]/div/input")
giris.send_keys("tracking key is mandatory")
giris.send_keys(Keys.ENTER)
time.sleep(2)

button_diff=browser.find_elements(By.CSS_SELECTOR,"span.msc-icon-checkmark.icon-negative-round-border")
button_elements = browser.find_elements(By.CSS_SELECTOR, "span.msc-icon-plus.icon-secondary")

containers = browser.find_elements(By.CSS_SELECTOR, ".data-value[x-text='container.ContainerNumber']")
containertype = browser.find_elements(By.CSS_SELECTOR, ".data-value[x-text='container.ContainerType']") 
latestmove = browser.find_elements(By.CSS_SELECTOR, ".data-value[x-text='container.LatestMove']")
podet = browser.find_elements(By.CSS_SELECTOR, ".data-value[x-text='container.PodEtaDate']")


CountFlag = 0
containerIndexList = []
containerIndex = 0

for index, button in enumerate(button_diff):
    button.click()
    
    # Tarih öğelerini al
    date_elements = browser.find_elements(By.CSS_SELECTOR, "span.data-value[x-text='event.Date']")
    valid_dates = [date_element.text for date_element in date_elements if date_element.text.strip() != ""]
    print(len(valid_dates))
    
    count = len(valid_dates) - CountFlag
    CountFlag = len(valid_dates)
    for i in range(count):
        containerIndexList.append(containers[containerIndex].text)
    
    containerIndex+=1
    
    # Diğer verileri al
    location = browser.find_elements(By.CSS_SELECTOR, "span.data-value[x-text='convertToSentenceCaseLocation(event.Location)']")
    valid_location = [location_element.text for location_element in location if location_element.text.strip() != ""]
    
    description = browser.find_elements(By.CSS_SELECTOR, "span.data-value[x-text='event.Description']")
    valid_description = [description_element.text for description_element in description if description_element.text.strip() != ""]
    
    tracking = browser.find_elements(By.CSS_SELECTOR, ".msc-flow-tracking_cell.msc-flow-trackingcell--five.msc-flow-tracking_cell--container")
    valid_tracking = [tracking_element.text for tracking_element in tracking if tracking_element.text.strip() != ""]
    
    equipment = browser.find_elements(By.CSS_SELECTOR, "span.data-value[x-text='convertToSentenceCase(event.EquipmentHandling.Name)']")
    valid_equipment = [equipment_element.text for equipment_element in equipment if equipment_element.text.strip() != ""]
    
   
   
workbook_path = "C:/Users/Ahmet Talha/Desktop/staj/proje/neden.xlsx"  # Kullanıcı adınıza göre düzenleyin
workbook = openpyxl.load_workbook(workbook_path)
    


sheet = workbook['Sheet']
sheet1 = workbook['Sheet1']

column_headers = ["Container", "Type", "Latest Move", "Pod Eta"]

for col_idx, header in enumerate(column_headers, start=1):
    sheet.cell(row=1, column=col_idx, value=header)
    

for index, container in enumerate(containers, start=2):
    sheet.cell(row=index, column=1, value=container.text)
for index, container in enumerate(containertype, start=2):
    sheet.cell(row=index, column=2, value=container.text)
for index, container in enumerate(latestmove, start=2):
    sheet.cell(row=index, column=3, value=container.text)
for index, container in enumerate(podet, start=2):
    sheet.cell(row=index, column=4, value=container.text)

    workbook.save(workbook_path)
    
column_headers2 = ["Container", "Date", "Actual/Estimated","Location", "Description", "Empty/Laden/Vessel/Voyage", "Equipment handling facility name"]
for qq, tr in enumerate(column_headers2, start=1): 
    sheet1.cell(row=1, column=qq, value=tr)

date_elements = browser.find_elements(By.CSS_SELECTOR, "span.data-value[x-text='event.Date']")
location = browser.find_elements(By.CSS_SELECTOR, "span.data-value[x-text='convertToSentenceCaseLocation(event.Location)']")
description = browser.find_elements(By.CSS_SELECTOR, "span.data-value[x-text='event.Description']")
tracking = browser.find_elements(By.CSS_SELECTOR, ".msc-flow-tracking__cell.msc-flow-tracking__cell--five.msc-flow-tracking__cell--container")
equipment = browser.find_elements(By.CSS_SELECTOR, "span.data-value[x-text='convertToSentenceCase(event.EquipmentHandling.Name)']")



for index, veriler in enumerate(containerIndexList, start=2):
    sheet1.cell(row=index, column=1, value=veriler)
for index, veriler in enumerate(date_elements, start=2):
    sheet1.cell(row=index, column=2, value=veriler.text)
for index, veriler in enumerate(location, start=2):
    sheet1.cell(row=index, column=4, value=veriler.text)
for index, veriler in enumerate(description, start=2):
    sheet1.cell(row=index, column=5, value=veriler.text)
for index, veriler in enumerate(tracking, start=2):
    sheet1.cell(row=index, column=6, value=veriler.text)
for index, veriler in enumerate(equipment, start=2):
    sheet1.cell(row=index, column=7, value=veriler.text)

    workbook.save(workbook_path)
    



    

